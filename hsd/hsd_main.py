import logging
import argparse
import requests
import logging
import urllib3
import json
import pdb
import time
import os
import openpyxl
from requests_kerberos import HTTPKerberosAuth
from openpyxl.styles import Font, PatternFill, Border


# Ignore the SSL insecure warning as we are passing in 'verify=false'
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logger = logging.getLogger('TCD')
logger.setLevel(logging.DEBUG)
handler = logging.StreamHandler()
formater = logging.Formatter('%(asctime)s - %(levelname)s: %(message)s')
handler.setFormatter(formater)
logger.addHandler(handler)

class Excel:
    def __init__(self):
        self.workbook = None
        self.sheet = None
        self.max_row = 0
        self.row = 0
        self.max_column = 0
        self.column = 0

    def open(self, srcfile):
        if not os.path.isfile(srcfile):
            logging.error("{ not exist!" % (srcfile))
            return

        openpyxl.Workbook.encoding = "utf8"
        self.file = srcfile
        self.workbook = openpyxl.load_workbook(filename=srcfile)
        self.sheet = self.workbook[self.workbook.sheetnames[0]]
        self.max_row = self.sheet.max_row
        self.row = 0
        self.max_column = self.sheet.max_column
        self.column = 0
        return

    def read(self):
        lines = []
        for row in self.sheet[1:self.max_row]:
            line = []
            for cell in row:
                if cell.value is None:
                    line.append('')
                else:
                    line.append(cell.value)

            lines.append(line)

        return lines

    def write(self, row, column, value):
        cell = self.sheet.cell(row=row, column=column, value=value)
        return

    def save_close(self):
        self.workbook.save(self.file)
        return


def get_parent_id(article_id):
    if not article_id:
        return 0

    headers = {'Content-type': 'application/json'}
    url = 'https://hsdes-api.intel.com/rest/article/' + article_id
    response = requests.get(url, verify=False, auth=HTTPKerberosAuth(), headers=headers)
    if response.status_code != 200:
        raise Exception("Invalid article_id!", article_id, response.status_code)

    parent_id = response.json()['data'][0]['from_id']
    logger.info('Article Parent ID(%s): %s', article_id, parent_id)
    return parent_id


def get_article_title(article_id):
    if not article_id or article_id == 0:
        return 'NA'

    headers = {'Content-type': 'application/json'}
    url = 'https://hsdes-api.intel.com/rest/article/' + article_id
    response = requests.get(url, verify=False, auth=HTTPKerberosAuth(), headers=headers)
    if response.status_code != 200:
        raise Exception("Invalid article_id!", article_id, response.status_code)

    title = response.json()['data'][0]['title']
    logger.info('Article Title(%s): %s', article_id, title)
    return title


def get_article_duration(article_id):
    if not article_id:
        return 0

    headers = {'Content-type': 'application/json'}
    url = 'https://hsdes-api.intel.com/rest/article/' + article_id
    response = requests.get(url, verify=False, auth=HTTPKerberosAuth(), headers=headers)
    if response.status_code != 200:
        raise Exception("Invalid article_id!", article_id, response.status_code)

    duration = response.json()['data'][0]['server_platf.test_case_definition.execution_runtime_effort']
    logger.info('Article Duration(%s): %s', article_id, duration)
    return duration

def get_article_links(article_id):
    if not article_id:
        return 0

    headers = {'Content-type': 'application/json'}
    url = 'https://hsdes-api.intel.com/rest/article/' + article_id + '/links'
    response = requests.get(url, verify=False, auth=HTTPKerberosAuth(), headers=headers)
    if response.status_code != 200:
        raise Exception("Invalid article_id!", article_id, response.status_code)

    return response.json()['responses']

def get_dell_id(article_id):
    if not article_id:
        return 0

    links = get_article_links(article_id)
    for link in links:
        if '[DELL]' in link['title'].upper():
            return link['id']
        elif '[CO-VAL]' in link['title'].upper():
            sub_links = get_article_links(link['id'])
            for sub_link in sub_links:
                if '[DELL]' in sub_link['title'].upper():
                    return sub_link['id']

    return 0

def get_coval_id(article_id):
    if not article_id:
        return 0

    links = get_article_links(article_id)
    for link in links:
        if '[CO-VAL]' in link['title'].upper():
            return link['id']

    return 0

def get_article_tag(article_id):
    if not article_id:
        return 0

    headers = {'Content-type': 'application/json'}
    url = 'https://hsdes-api.intel.com/rest/article/' + article_id + '?fields=tag'
    response = requests.get(url, verify=False, auth=HTTPKerberosAuth(), headers=headers)
    if response.status_code != 200:
        raise Exception("Invalid article_id!", article_id, response.status_code)

    return response.json()['data'][0]['tag']

def update_article_tag(article_id, updated_tag):
    if not article_id:
        return 0

    old_tag = get_article_tag(article_id)
    logger.info("Old Tag of HSD#{}: {}".format(article_id, old_tag))

    headers = {'Content-type': 'application/json'}
    url = 'https://hsdes-api.intel.com/rest/article/' + article_id
    data = {
        "tenant": "server_platf",
        "subject": "test_case_definition",
        "fieldValues": [
            {
                "tag": old_tag + "," + updated_tag
            },
            {
                "send_mail": "false"
            }
        ]
    }
    response = requests.put(url, verify=False, auth=HTTPKerberosAuth(), headers=headers, data=json.dumps(data))
    if response.status_code != 200:
        raise Exception("Failed to update Article tag!", article_id, response.status_code)

    new_tag = get_article_tag(article_id)
    logger.info("New Tag of HSD#{}: {}".format(article_id, new_tag))

def update_owner_team(article_id):
    headers = {'Content-type': 'application/json'}
    url = 'https://hsdes-api.intel.com/rest/article/' + article_id
    data = {
        "tenant": "server_platf",
        "subject": "test_case_definition",
        "fieldValues": [
            {
                "test_case_definition.owner_team": "PAIV.Coval"
            },
            {
                "send_mail": "false"
            }
        ]
    }
    response = requests.put(url, verify=False, auth=HTTPKerberosAuth(), headers=headers, data=json.dumps(data))
    if response.status_code != 200:
        raise Exception("Failed to update Article owner_team!", article_id, response.status_code)

def get_parent_priority(article_id):
    headers = {'Content-type': 'application/json'}
    url = 'https://hsdes-api.intel.com/rest/article/' + article_id
    response = requests.get(url, verify=False, auth=HTTPKerberosAuth(), headers=headers)
    if response.status_code != 200:
        raise Exception("Invalid article_id!", article_id, response.status_code)

    parent_article_id = response.json()['data'][0]['from_id']
    logger.info('Parent of Article(%s): %s', article_id, parent_article_id)
    parent_url = 'https://hsdes-api.intel.com/rest/article/' + parent_article_id
    response = requests.get(parent_url, verify=False, auth=HTTPKerberosAuth(), headers=headers)
    if response.status_code != 200:
        raise Exception("Invalid article_id!", article_id, response.status_code)

    parent_priority = response.json()['data'][0]['priority']
    logger.info('Priority of Parent Article(%s): %s', parent_article_id, parent_priority)
    
    return parent_priority

def update_priority(article_id):
    headers = {'Content-type': 'application/json'}
    url = 'https://hsdes-api.intel.com/rest/article/' + article_id
    priority = get_parent_priority(article_id)
    data = {
        "tenant": "server_platf",
        "subject": "test_case_definition",
        "fieldValues": [
            {
                "priority": priority
            },
            {
                "send_mail": "false"
            }
        ]
    }
    response = requests.put(url, verify=False, auth=HTTPKerberosAuth(), headers=headers, data=json.dumps(data))
    if response.status_code != 200:
        raise Exception("Failed to update Article priority!", article_id, response.status_code)

def clone_article(article_id):
    title = get_article_title(article_id)
    tags = get_article_tag(article_id)

    headers = {'Content-type': 'application/json'}
    url = 'https://hsdes-api.intel.com/rest/article/' + article_id + '/clone'
    
    data = {
        "destTenant": "server_platf",
        "destSubject": "test_case_definition",
        "sendmail": "false",
        "copy_attachment": "true",
        "copy_comment": "false",
        "fieldValues": [
            {
                "title": "[DELL] " + title
            },
            {
                "tag": tags + ", DELL_GNRSP_YES, GNRSP_QS_PC"
            },
            {
                "test_case_definition.free_tag_1": article_id
            },
        ]
    }
    response = requests.post(url, verify=False, auth=HTTPKerberosAuth(), headers=headers, data=json.dumps(data))

    if response.status_code != 200:
        raise Exception("Article clone failed!", article_id, response.status_code)
    else:
        logger.info('Article(%s) was cloned to new ID %s successfully. Please refer to: https://hsdes.intel.com/appstore/article/#%s', article_id, response.json()['newID'], response.json()['newID'])
        time.sleep(5)
        # WA to solve the owner_team could not be updated to "PAIV.Coval" during article clone
        update_owner_team(str(response.json()['newID']))
        # WA to solve the priority could not be could not be inherited during article clone
        update_priority(str(response.json()['newID']))
        return response.json()['newID']

def get_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--example", type=str, default='', help="Just an example!")
    args = parser.parse_args()

    return args


if __name__ == '__main__':
    args = get_args()
    excel = Excel()
    excel.open('TCD.xlsx')

    lines = excel.read()
    logger.info("To Captured TCD Number: {}".format(str(len(lines))))
    try:
        for line in range(1,len(lines) + 1):
            if line == 1:
                #excel.write(line, 2, 'rp_tcd_title')
                #excel.write(line, 3, 'dell_tcd_id')
                #excel.write(line, 4, 'dell_tcd_title')
                #excel.write(line, 5, 'coval_tcd_id')
                #excel.write(line, 6, 'coval_tcd_title')
                continue
            #update_article_tag(str(lines[line-1][0]), "From_GNRSP_BKC")
            clone_article(str(lines[line-1][0]))
            '''
            rp_tcd_id = str(lines[line-1][0])
            logger.info("Starting to capture TCD via Line#{}: {}".format(line, rp_tcd_id))
            rp_tcd_title = get_article_title(rp_tcd_id)
            dell_tcd_id = get_dell_id(rp_tcd_id)
            dell_tcd_title = get_article_title(dell_tcd_id)
            #coval_tcd_id = get_coval_id(rp_tcd_id)
            #time.sleep(1)
            #coval_tcd_title = get_article_title(coval_tcd_id)

            excel.write(line, 2, rp_tcd_title)
            excel.write(line, 3, dell_tcd_id)
            excel.write(line, 4, dell_tcd_title)
            #excel.write(line, 5, coval_tcd_id)
            #excel.write(line, 6, coval_tcd_title)
            '''
    except Exception as e:
        print(e)
        excel.save_close()

    excel.save_close()